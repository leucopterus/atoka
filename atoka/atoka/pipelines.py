# Define your item pipelines here
#
# Don't forget to add your pipeline to the ITEM_PIPELINES setting
# See: https://docs.scrapy.org/en/latest/topics/item-pipeline.html


# useful for handling different item types with a single interface
import os

import openpyxl
from openpyxl.styles import Font
from scrapy.exceptions import DropItem

from .items import (
    AtokaContactsItem,
    AtokaErrorContactsItem,
)
from .settings import BASE_DIR


class ExcelOutputPipeline:
    def __init__(self):
        self.wb_path = os.path.join(BASE_DIR, 'atoka/spiders/output/output2000-2.xlsx')
        self.error_wb_path = os.path.join(BASE_DIR, 'atoka/spiders/output/error2000-2.xlsx')
        self.cod_fiscale_row_mapping = {}
        self.last_row_output = 1
        output = [
            'COMPANY_NAME',
            'TAX_CODE',
            'VAT_NUMBER',
            'EMAIL',
            'PHONE',
            'WEBSITE',
        ]

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(output)
        row = ws.row_dimensions[1]
        row.font = Font(bold=True)
        wb.save(self.wb_path)
        wb.close()

    def process_item(self, item, spider):
        if isinstance(item, AtokaContactsItem):
            self._fill_excel_with_company_data(item)
            return item
        elif isinstance(item, AtokaErrorContactsItem):
            self._fill_error_excel_with_code_data(item)
        raise DropItem

    def _collect_items_from_list(self, objects, field):
        return [obj.get(f'{field}') for obj in objects]

    def _fill_excel_with_company_data(self, data=None):
        if data is None:
            return

        code = data.get('code')
        company_name = data.get('company_name')
        vat_id = data.get('vat_id')
        all_emails = self._collect_items_from_list(data.get('emails'), 'address')
        all_phones = self._collect_items_from_list(data.get('phones'), 'number')
        all_websites = self._collect_items_from_list(data.get('websites'), 'url')
        max_items_number = max(len(all_emails), len(all_phones), len(all_websites), 1)

        wb = openpyxl.load_workbook(self.wb_path)
        ws = wb.active
        for i in range(max_items_number):
            email = all_emails[i:i+1][0] if all_emails[i:i+1] else None
            phone = all_phones[i:i+1][0] if all_phones[i:i+1] else None
            website = all_websites[i:i+1][0] if all_websites[i:i+1] else None
            output = [company_name, code, vat_id, email, phone, website]
            ws.append(output)
        wb.save(self.wb_path)
        wb.close()

    def _fill_error_excel_with_code_data(self, item):
        if not os.path.isfile(self.error_wb_path):
            wb = openpyxl.Workbook()
        else:
            wb = openpyxl.load_workbook(self.error_wb_path)
        ws = wb.active
        ws.append([item.get('code'), item.get('reason')])
        wb.save(self.error_wb_path)
        wb.close()
